from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
SUMMARY_PATH = ROOT / "results" / "summary.csv"
SWEEP_PATH = ROOT / "results" / "sweep_summary.csv"
REPEATED_PATH = ROOT / "results" / "repeated_summary.csv"
GRID_PATH = ROOT / "results" / "grid_summary.csv"
SCENARIO_PATH = ROOT / "results" / "scenario_summary.csv"
REPORT_PATH = ROOT / "results" / "edge_cache_fallback_report.xlsx"


def main() -> None:
    if not SUMMARY_PATH.exists():
        raise FileNotFoundError("Run scripts/run_experiment.py before building the report.")

    data = _read_csv(SUMMARY_PATH)
    sweep_data = _read_csv(SWEEP_PATH) if SWEEP_PATH.exists() else None
    repeated_data = _read_csv(REPEATED_PATH) if REPEATED_PATH.exists() else None
    grid_data = _read_csv(GRID_PATH) if GRID_PATH.exists() else None
    scenario_data = _read_csv(SCENARIO_PATH) if SCENARIO_PATH.exists() else None

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    parameters = workbook.create_sheet("Parameters")
    summary = workbook.create_sheet("Summary")
    charts = workbook.create_sheet("Charts")
    origin_sweep = workbook.create_sheet("Origin Delay Sweep")
    availability_sweep = workbook.create_sheet("ES Availability Sweep")
    b2_advantage = workbook.create_sheet("B2 Advantage")
    formal_scenarios = workbook.create_sheet("Formal Scenarios")
    repeated_trials = workbook.create_sheet("Repeated Trials")
    b2_grid = workbook.create_sheet("B2 Advantage Grid")

    _build_overview(overview, data, sweep_data, repeated_data, grid_data, scenario_data)
    _build_parameters(parameters, data)
    _build_summary(summary, data)
    _build_charts(charts, data)
    if sweep_data is not None:
        _build_sweep_sheet(
            origin_sweep,
            [row for row in sweep_data if row["sweep_name"] == "origin_delay"],
            "origin_delay",
            "Origin Delay Sweep",
        )
        _build_sweep_sheet(
            availability_sweep,
            [row for row in sweep_data if row["sweep_name"] == "es_availability"],
            "es_availability",
            "ES Availability Sweep",
        )
        _build_b2_advantage(b2_advantage, sweep_data)
    else:
        _build_empty_note(origin_sweep, "Run scripts/run_sweep.py, then rerun this report.")
        _build_empty_note(availability_sweep, "Run scripts/run_sweep.py, then rerun this report.")
        _build_empty_note(b2_advantage, "Run scripts/run_sweep.py, then rerun this report.")

    if repeated_data is not None:
        _build_repeated_trials(repeated_trials, repeated_data)
    else:
        _build_empty_note(repeated_trials, "Run scripts/run_repeated.py to add CI results.")

    if scenario_data is not None:
        _build_formal_scenarios(formal_scenarios, scenario_data)
    else:
        _build_empty_note(formal_scenarios, "Run scripts/run_scenarios.py to add scenario results.")

    if grid_data is not None:
        _build_b2_grid(b2_grid, grid_data)
    else:
        _build_empty_note(b2_grid, "Run scripts/run_repeated.py to add the heatmap grid.")

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        _size_columns(sheet)

    REPORT_PATH.parent.mkdir(exist_ok=True)
    workbook.save(REPORT_PATH)
    print(f"Wrote {REPORT_PATH}")


def _read_csv(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _build_overview(
    sheet,
    data: list[dict],
    sweep_data: list[dict] | None,
    repeated_data: list[dict] | None,
    grid_data: list[dict] | None,
    scenario_data: list[dict] | None,
) -> None:
    best_mean = min(data, key=lambda row: _float(row["mean_response_time"]))
    best_p95 = min(data, key=lambda row: _float(row["p95_response_time"]))
    sweep_note = (
        "Sweep results included: origin delay and ES availability."
        if sweep_data is not None
        else "No sweep results found yet. Run scripts/run_sweep.py to add them."
    )
    repeated_note = (
        "Repeated-trial CI results included."
        if repeated_data is not None
        else "No repeated-trial results found yet. Run scripts/run_repeated.py to add them."
    )
    grid_note = (
        "B2 advantage grid included."
        if grid_data is not None
        else "No grid results found yet. Run scripts/run_repeated.py to add them."
    )
    scenario_note = (
        "Formal three-scenario results included."
        if scenario_data is not None
        else "No formal scenario results found yet. Run scripts/run_scenarios.py to add them."
    )

    rows = [
        ["Edge Cache Fallback Report", ""],
        ["Scenario", data[0]["scenario"]],
        ["Policies", ", ".join(row["policy"] for row in data)],
        [
            "Best mean response time",
            f"{best_mean['policy']} ({_float(best_mean['mean_response_time']):.3f})",
        ],
        ["Best p95 response time", f"{best_p95['policy']} ({_float(best_p95['p95_response_time']):.3f})"],
        ["Main use", "Readable first-stage report; CSV remains the reproducible data source."],
        ["Sweep status", sweep_note],
        ["Repeated-trial status", repeated_note],
        ["Grid status", grid_note],
        ["Scenario status", scenario_note],
    ]
    for row in rows:
        sheet.append(row)

    sheet.merge_cells("A1:B1")
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    for cell in sheet["A"]:
        cell.font = Font(bold=True)
    for row_index in range(6, 11):
        sheet[f"A{row_index}"].alignment = Alignment(wrap_text=True)
        sheet[f"B{row_index}"].alignment = Alignment(wrap_text=True)


def _build_parameters(sheet, data: list[dict]) -> None:
    parameter_cols = [
        "zipf_alpha",
        "es_availability",
        "local_es_availability",
        "neighbor_es_availability",
        "origin_delay",
        "local_es_count",
        "neighbor_group_size",
        "k",
    ]
    sheet.append(["Parameter", "Value"])
    first_row = data[0]
    for name in parameter_cols:
        sheet.append([name, _coerce(first_row.get(name, ""))])
    _style_header(sheet, "A1:B1")
    _add_table(sheet, f"A1:B{len(parameter_cols) + 1}", "ParametersTable")


def _build_summary(sheet, data: list[dict]) -> None:
    display_cols = [
        "policy",
        "mean_response_time",
        "p95_response_time",
        "origin_free_rate",
        "neighbor_failure_rate",
    ]
    sheet.append(display_cols)
    for row in data:
        sheet.append([_coerce(row[column]) for column in display_cols])

    _style_header(sheet, "A1:E1")
    _add_table(sheet, f"A1:E{len(data) + 1}", "SummaryTable")

    for col in ("B", "C"):
        for cell in sheet[col][1:]:
            cell.number_format = "0.000"
    for col in ("D", "E"):
        for cell in sheet[col][1:]:
            cell.number_format = "0.00%"


def _build_charts(sheet, data: list[dict]) -> None:
    chart_cols = [
        "policy",
        "mean_response_time",
        "p95_response_time",
        "origin_free_rate",
        "neighbor_failure_rate",
    ]
    sheet.append(chart_cols)
    for row in data:
        sheet.append([_coerce(row[column]) for column in chart_cols])

    _style_header(sheet, "A1:E1")

    _add_chart(sheet, "Mean Response Time", 2, "G2")
    _add_chart(sheet, "P95 Response Time", 3, "G18")
    _add_chart(sheet, "Origin-Free Rate", 4, "A18")
    _add_chart(sheet, "Neighbor Failure Rate", 5, "A34")


def _build_sweep_sheet(sheet, data: list[dict], sweep_name: str, title: str) -> None:
    if not data:
        _build_empty_note(sheet, "No rows found for this sweep.")
        return

    sheet.append([title])
    sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    sheet.merge_cells("A1:F1")

    metrics = [
        "mean_response_time",
        "p95_response_time",
        "origin_free_rate",
        "neighbor_failure_rate",
    ]
    start_row = 3
    for metric in metrics:
        pivot = _pivot_policy_metric(data, metric)
        sheet.append([sweep_name, "B0", "B1", "B2"])
        for sweep_value, policy_values in pivot:
            sheet.append([sweep_value, policy_values.get("B0"), policy_values.get("B1"), policy_values.get("B2")])
        _style_header(sheet, f"A{start_row}:D{start_row}")
        _add_line_chart(
            sheet,
            title=metric,
            min_row=start_row,
            max_row=start_row + len(pivot),
            anchor=f"F{start_row}",
            percent=("rate" in metric),
        )
        start_row += len(pivot) + 17


def _build_b2_advantage(sheet, sweep_data: list[dict]) -> None:
    sheet.append(["B2 Advantage vs B1"])
    sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    sheet.merge_cells("A1:E1")

    rows = [["sweep_name", "sweep_value", "b1_mean", "b2_mean", "b2_advantage_vs_b1"]]
    for (sweep_name, sweep_value), group in _group_by(sweep_data, ["sweep_name", "sweep_value"]).items():
        by_policy = {row["policy"]: row for row in group}
        b1_mean = _float(by_policy["B1"]["mean_response_time"])
        b2_mean = _float(by_policy["B2"]["mean_response_time"])
        rows.append([sweep_name, _coerce(sweep_value), b1_mean, b2_mean, round(b1_mean - b2_mean, 3)])

    for row in rows:
        sheet.append(row)
    _style_header(sheet, "A2:E2")
    _add_table(sheet, f"A2:E{len(rows) + 1}", "B2AdvantageTable")

    chart = BarChart()
    chart.title = "B2 Mean Response-Time Advantage vs B1"
    chart.y_axis.title = "B1 mean - B2 mean"
    chart.x_axis.title = "Sweep case"
    chart.height = 8
    chart.width = 16
    data = Reference(sheet, min_col=5, min_row=2, max_row=len(rows) + 1)
    cats = Reference(sheet, min_col=2, min_row=3, max_row=len(rows) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    sheet.add_chart(chart, "G2")


def _build_repeated_trials(sheet, data: list[dict]) -> None:
    sheet.append(["Repeated-Trial Summary"])
    sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    sheet.merge_cells("A1:L1")

    display_cols = [
        "scenario",
        "sweep_name",
        "sweep_value",
        "policy",
        "trial_count",
        "mean_response_time_mean",
        "mean_response_time_ci95_low",
        "mean_response_time_ci95_high",
        "p95_response_time_mean",
        "origin_free_rate_mean",
        "neighbor_failure_rate_mean",
        "b2_advantage_vs_b1_mean",
    ]
    sheet.append(display_cols)
    for row in data:
        sheet.append([_coerce(row.get(column, "")) for column in display_cols])

    _style_header(sheet, "A2:L2")
    _add_table(sheet, f"A2:L{len(data) + 2}", "RepeatedSummaryTable")

    for col in ("F", "G", "H", "I", "L"):
        for cell in sheet[col][2:]:
            cell.number_format = "0.000"
    for col in ("J", "K"):
        for cell in sheet[col][2:]:
            cell.number_format = "0.00%"


def _build_formal_scenarios(sheet, data: list[dict]) -> None:
    sheet.append(["Formal Three-Scenario Summary"])
    sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    sheet.merge_cells("A1:M1")

    display_cols = [
        "scenario",
        "policy",
        "trial_count",
        "origin_delay",
        "local_es_availability",
        "neighbor_es_availability",
        "mean_response_time_mean",
        "mean_response_time_ci95_low",
        "mean_response_time_ci95_high",
        "p95_response_time_mean",
        "origin_free_rate_mean",
        "neighbor_failure_rate_mean",
        "b2_advantage_vs_b1_mean",
    ]
    sheet.append(display_cols)
    for row in data:
        sheet.append([_coerce(row.get(column, "")) for column in display_cols])

    _style_header(sheet, "A2:M2")
    _add_table(sheet, f"A2:M{len(data) + 2}", "FormalScenariosTable")

    for col in ("D", "G", "H", "I", "J", "M"):
        for cell in sheet[col][2:]:
            cell.number_format = "0.000"
    for col in ("E", "F", "K", "L"):
        for cell in sheet[col][2:]:
            cell.number_format = "0.00%"


def _build_b2_grid(sheet, grid_data: list[dict]) -> None:
    b2_rows = [row for row in grid_data if row["policy"] == "B2"]
    if not b2_rows:
        _build_empty_note(sheet, "No B2 rows found in grid_summary.csv.")
        return

    origin_delays = sorted({_float(row["origin_delay"]) for row in b2_rows})
    availabilities = sorted({_float(row["es_availability"]) for row in b2_rows})
    lookup = {
        (_float(row["es_availability"]), _float(row["origin_delay"])): _float(row["b2_advantage_vs_b1_mean"])
        for row in b2_rows
        if row.get("b2_advantage_vs_b1_mean", "") != ""
    }

    sheet.append(["B2 Advantage Grid"])
    sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(origin_delays) + 1)

    sheet.append(["es_availability / origin_delay", *origin_delays])
    for availability in availabilities:
        sheet.append([availability, *[lookup.get((availability, origin_delay), "") for origin_delay in origin_delays]])

    _style_header(sheet, f"A2:{get_column_letter(len(origin_delays) + 1)}2")

    for row in sheet.iter_rows(
        min_row=3,
        max_row=2 + len(availabilities),
        min_col=2,
        max_col=len(origin_delays) + 1,
    ):
        for cell in row:
            if cell.value == "":
                continue
            cell.number_format = "0.000"
            if float(cell.value) > 0.5:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
            elif float(cell.value) < -0.5:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
            else:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")


def _build_empty_note(sheet, instruction: str) -> None:
    sheet.append(["Results not available"])
    sheet.append([instruction])
    sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E79")


def _pivot_policy_metric(rows: list[dict], metric: str) -> list[tuple[float, dict[str, float]]]:
    grouped: dict[float, dict[str, float]] = {}
    for row in rows:
        sweep_value = _float(row["sweep_value"])
        grouped.setdefault(sweep_value, {})[row["policy"]] = _float(row[metric])
    return sorted(grouped.items())


def _group_by(rows: list[dict], columns: list[str]) -> dict[tuple, list[dict]]:
    grouped: dict[tuple, list[dict]] = {}
    for row in rows:
        key = tuple(row[column] for column in columns)
        grouped.setdefault(key, []).append(row)
    return grouped


def _add_line_chart(
    sheet,
    title: str,
    min_row: int,
    max_row: int,
    anchor: str,
    percent: bool = False,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = title
    chart.x_axis.title = "Sweep value"
    chart.height = 7
    chart.width = 12
    values = Reference(sheet, min_col=2, max_col=4, min_row=min_row, max_row=max_row)
    categories = Reference(sheet, min_col=1, min_row=min_row + 1, max_row=max_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    if percent:
        chart.y_axis.numFmt = "0%"
    sheet.add_chart(chart, anchor)


def _add_chart(sheet, title: str, value_col: int, anchor: str) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = title
    chart.x_axis.title = "Policy"
    chart.height = 7
    chart.width = 12

    values = Reference(sheet, min_col=value_col, min_row=1, max_row=4)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=4)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    sheet.add_chart(chart, anchor)


def _style_header(sheet, range_name: str) -> None:
    for row in sheet[range_name]:
        for cell in row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")


def _add_table(sheet, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _size_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 34)


def _coerce(value):
    if value is None or value == "":
        return ""
    text = str(value)
    try:
        number = float(text)
    except ValueError:
        return value
    if number.is_integer():
        return int(number)
    return number


def _float(value) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


if __name__ == "__main__":
    main()
